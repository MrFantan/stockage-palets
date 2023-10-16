import openpyxl
from openpyxl import Workbook
import pandas as pd
from datetime import datetime

def ingresar_datos():
    referencia = input("Código de referencia: F")
    cliente = input("Nombre cliente: ")
    unidades = input("Unidades por embalaje: ")
    embalajes = input("Cantidad de cajas en el palet: ")
    palets = int(input("Cantidad de palets: "))
    
    # Validar que la cantidad de palets sea al menos 1
    if palets < 1:
        print("Error: La cantidad de palets debe ser al menos 1.")
        return

    # Crear una lista para almacenar los datos de cada palet
    palets_data = []

    # Determinar el índice inicial (la última fila + 1)
    try:
        workbook = openpyxl.load_workbook('./Datos/Basedatos/Almacen.xlsx')
        sheet = workbook.active
        indice_inicial = sheet.max_row + 1
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Índice", "Referencia", "Cliente", "Unidades", "Embalajes", "Cantidad Total", "Palets", "Calle", "Base", "Altura", "Trabajador", "Fecha", "Hora"])
        indice_inicial = 1

    # Ingresar datos para cada palet
    for palet_num in range(1, palets + 1):
        while True:
            calle = input(f"En qué calle está ubicado el palet {palet_num} (A-F): ")
            base = input(f"En qué base está ubicado el palet {palet_num} (1-33): ")
            altura = input(f"En qué altura está ubicado el palet {palet_num} (0-8): ")

            # Validar que las ubicaciones sean correctas
            if calle.upper() in ['A', 'B', 'C', 'D', 'E', 'F'] and base.isdigit() and 1 <= int(base) <= 33 and altura.isdigit() and 0 <= int(altura) <= 8:
                break
            else:
                print("Ubicación incorrecta. Asegúrate de que la calle esté entre A y F, la base esté entre 1 y 33, y la altura entre 0 y 8.")

        trabajador = input("Cuál es su nombre: ")

        # Formatear fecha y hora
        ahora = datetime.today()
        dia = ahora.strftime("%d/%m/%y")
        hora = ahora.strftime("%H:%M")

        datos = {
            "Índice": indice_inicial,
            "Referencia": "F" + referencia,
            "Cliente": cliente.upper(),
            "Unidades": unidades, 
            "Embalajes": embalajes,
            "Cantidad Total": int(unidades) * int(embalajes),
            "Palets": palet_num,
            "Calle": calle.upper(),
            "Base": base,
            "Altura": altura,
            "Trabajador": trabajador.title(),
            "Fecha": dia,
            "Hora": hora,
        }
        palets_data.append(datos)

        # Incrementar el índice para la siguiente fila
        indice_inicial += 1

    # Ingresar datos para cada palet
    for datos in palets_data:
        values = list(datos.values())
        sheet.append(values)

    workbook.save('./Datos/Basedatos/Almacen.xlsx')
    print(f"Se han agregado {palets} palets del cliente {cliente} con la referencia número: F{referencia}, cada palet con un total de {unidades} unidades por embalaje, distribuidas en {embalajes} cajas. Los palets se han colocado en las siguientes ubicaciones:")
    for datos in palets_data:
        print(f" - Calle: {datos['Calle']}, Base: {datos['Base']}, Altura: {datos['Altura']} por el trabajador: {datos['Trabajador']}, el día {datos['Fecha']} a las {datos['Hora']}.")
    print("Productos almacenados exitosamente.")
    
# Llama a la función ingresar_datos para agregar datos
ingresar_datos()
